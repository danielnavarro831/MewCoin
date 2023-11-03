import openpyxl


class DocReader:
    # location = './/file_name' on replit
    # The locations that start with 'C:' are for testing locally. When on the live server, comment these out and
    # switch to the './/' locations instead
    _strings_location = 'C:\\Users\\Daniel\\PycharmProjects\\MewCoinV3\\Data\\mewcoin_string_data_v1.xlsx'
    _commands_location = 'C:\\Users\\Daniel\\PycharmProjects\\MewCoinV3\\Data\\mewcoin_command_data_v1.xlsx'
    _variables_location = 'C:\\Users\\Daniel\\PycharmProjects\\MewCoinV3\\Data\\mewcoin_variable_data_v1.xlsx'
    _export_location = 'C:\\Users\\Daniel\\PycharmProjects\\MewCoinV3\\Data\\mewcoin_export_data_v1.xlsx'
    # _strings_location = './/Data//mewcoin_string_data_v1.xlsx'
    # _commands_location = './/Data//mewcoin_command_data_v1.xlsx'
    # _variables_location = './/Data//mewcoin_variable_data_v1.xlsx'
    # _export_location = './/Data//mewcoin_export_data_v1.xlsx'

    @classmethod
    def get_string(cls, string: str, args: dict):  # string is the name of the string, args is a dict of the words that
        # replace the variables like "UserName"
        book = openpyxl.load_workbook(DocReader._strings_location)  # Opens the excel file
        sheet = book["Strings"]  # gets the tab titled "Strings" in the opened excel file
        message = ""  # initializes a blank message
        for a in range(2, sheet.max_row + 1):  # Goes through each row of strings and tries to find a matching name
            if sheet.cell(row=a, column=1).value == string:
                message = sheet.cell(row=a, column=2).value  # sets "message" to the text in the string
                break
        if message:  # if a match was found, replace the variable words with the words given
            for a in args.keys():
                message = message.replace(a, str(args[a]), message.count(a))
            return message  # return the message to post a response
        message = "String Error - No Reference: " + string + " " + str(args)
        # If no message was found, return the String Error
        return message

    @classmethod
    def get_variable_data(cls, variable: str):  # variable is the name of the variable in the excel sheet
        book = openpyxl.load_workbook(DocReader._variables_location)  # opens the excel sheet
        sheet = book["Variables"]  # finds the tab titled "Variables" in the opened excel sheet
        for a in range(2, sheet.max_row + 1):  # goes through the rows and columns to find the matching name and value
            var_name = sheet.cell(row=a, column=1).value
            if var_name == variable:
                value = sheet.cell(row=a, column=2).value
                return value

    @classmethod
    def set_variable_data(cls, variable: str, value):
        book = openpyxl.load_workbook(DocReader._variables_location)
        sheet = book["Variables"]
        for a in range(2, sheet.max_row + 1):
            var_name = sheet.cell(row=a, column=1).value
            if var_name == variable:
                sheet.cell(row=a, column=2).value = value
                book.save(DocReader._variables_location)  # Sets the variable value to the value given
                break

    @classmethod
    def get_commands(cls):
        book = openpyxl.load_workbook(DocReader._commands_location)
        sheet = book["Commands"]
        command_list = {}
        for a in range(2, sheet.max_row + 1):
            command_dict = {}
            for b in range(1, sheet.max_column + 1):
                key = sheet.cell(row=1, column=b).value
                value = DocReader.interpret_cell_data(sheet.cell(row=a, column=b).value)
                if type(value) == str:
                    value = value.strip("'")
                command_dict[key] = value
            command_name = command_dict["Command"]
            command_list[command_name] = command_dict
        return command_list

    @classmethod
    def get_admin_commands(cls):
        book = openpyxl.load_workbook(DocReader._commands_location)
        sheet = book["Commands"]
        admin_command_list = {}
        admin_column = 0
        for a in range(2, sheet.max_row + 1):
            command_dict = {}
            for b in range(1, sheet.max_column + 1):
                key = sheet.cell(row=1, column=b).value
                if key == "Admin Command":
                    admin_column = b
                value = DocReader.interpret_cell_data(sheet.cell(row=a, column=b).value)
                if type(value) == str:
                    value = value.strip("'")
                command_dict[key] = value
            if sheet.cell(row=a, column=admin_column).value:
                command_name = command_dict["Command"]
                admin_command_list[command_name] = command_dict
        return admin_command_list

    @classmethod
    def interpret_cell_data(cls, value):
        if value:
            if value == "TRUE":
                return True
            else:
                return value
        else:
            return None

    @classmethod
    def export_mewcoin_data(cls, accounts: dict):
        book = openpyxl.load_workbook(DocReader._export_location)
        sheet = book["Balances"]
        line = 2
        for a in accounts.keys():
            account = str(a)
            balance = str(accounts[a])
            sheet.cell(row=line, column=1).value = account
            sheet.cell(row=line, column=2).value = balance
            line += 1
        book.save(DocReader._export_location)
